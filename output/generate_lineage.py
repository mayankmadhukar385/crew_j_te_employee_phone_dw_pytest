"""
Generate lineage Excel and Mermaid diagram for CREW_J_TE_EMPLOYEE_PHONE_DW.
Parses the DataStage XML export and produces:
  - output/CREW_J_TE_EMPLOYEE_PHONE_DW_lineage.xlsx  (3-tab workbook)
  - output/CREW_J_TE_EMPLOYEE_PHONE_DW_lineage_diagram.md
"""

import os
import re
import xml.etree.ElementTree as ET
from collections import defaultdict, deque

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required: pip install openpyxl")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XML_PATH = os.path.join(BASE_DIR, "input", "CREW_J_TE_EMPLOYEE_PHONE_DW.xml")
OUT_DIR  = os.path.join(BASE_DIR, "output")
XLSX_OUT = os.path.join(OUT_DIR, "CREW_J_TE_EMPLOYEE_PHONE_DW_lineage.xlsx")
MD_OUT   = os.path.join(OUT_DIR, "CREW_J_TE_EMPLOYEE_PHONE_DW_lineage_diagram.md")

JOB_NAME = "CREW_J_TE_EMPLOYEE_PHONE_DW"

# ---------------------------------------------------------------------------
# Stage graph - hand-built from XML metadata we extracted
# The XML lines 142-158 give us the definitive stage/link maps.
# ---------------------------------------------------------------------------

# Definitive stage list from IdentList (line 87 / StageNames line 142)
STAGE_NAMES_ORDERED = [
    "WSTELE_LND_TDC",
    "SEPARATE_PHTYPE_TFM",
    "BASIC_TYPE_PVT",
    "BASIC_REC_TFM",
    "HOME_SEQ1_PVT",
    "HOME_CPY",
    "HOME_SEQ2_PVT",
    "HOME_SEQ3_PVT",
    "HOME_SEQ_FNL",
    "INC_PRTY_TFM",
    "HOME_BASIC_JNR",
    "NEW_PRTY_TFM",
    "HOME_SEQ2_TFM",
    "HOME_SEQ3_TFM",
    "MAX_PRTY_AGG",
    "BASIC_MAX_JNR",
    "ADJUSTING_PRTY_TFM",
    "ALL_REC_FNL",
    "AWAY_SEQ3_TFM",
    "AWAY_SEQ2_TFM",
    "AWAY_SEQ_FNL",
    "AWAY_SEQ3_PVT",
    "AWAY_SEQ2_PVT",
    "AWAY_CPY",
    "AWAY_SEQ1_PVT",
    "INC_AWAY_PRTY_TFM",
    "AWAY_BASIC_JNR",
    "MAX_ABPRTY_AGG",
    "ABREC_NEW_PRTY_TFM",
    "AWAY_MAX_JNR",
    "ADJUST_PRTY_TFM",
    "TE_EMPLOYEE_PHONE_DW",
    "NULL_VAL_TFM",
    "WSSEN_ERR_SEQ",
]

# Stage type IDs from line 143
STAGE_TYPE_IDS = {
    "WSTELE_LND_TDC":      "TeradataConnectorPX",
    "SEPARATE_PHTYPE_TFM": "CTransformerStage",
    "BASIC_TYPE_PVT":      "PxPivot",
    "BASIC_REC_TFM":       "CTransformerStage",
    "HOME_SEQ1_PVT":       "PxPivot",
    "HOME_CPY":            "PxCopy",
    "HOME_SEQ2_PVT":       "PxPivot",
    "HOME_SEQ3_PVT":       "PxPivot",
    "HOME_SEQ_FNL":        "PxFunnel",
    "INC_PRTY_TFM":        "CTransformerStage",
    "HOME_BASIC_JNR":      "PxJoin",
    "NEW_PRTY_TFM":        "CTransformerStage",
    "HOME_SEQ2_TFM":       "CTransformerStage",
    "HOME_SEQ3_TFM":       "CTransformerStage",
    "MAX_PRTY_AGG":        "PxAggregator",
    "BASIC_MAX_JNR":       "PxJoin",
    "ADJUSTING_PRTY_TFM":  "CTransformerStage",
    "ALL_REC_FNL":         "PxFunnel",
    "AWAY_SEQ3_TFM":       "CTransformerStage",
    "AWAY_SEQ2_TFM":       "CTransformerStage",
    "AWAY_SEQ_FNL":        "PxFunnel",
    "AWAY_SEQ3_PVT":       "PxPivot",
    "AWAY_SEQ2_PVT":       "PxPivot",
    "AWAY_CPY":            "PxCopy",
    "AWAY_SEQ1_PVT":       "PxPivot",
    "INC_AWAY_PRTY_TFM":   "CTransformerStage",
    "AWAY_BASIC_JNR":      "PxJoin",
    "MAX_ABPRTY_AGG":      "PxAggregator",
    "ABREC_NEW_PRTY_TFM":  "CTransformerStage",
    "AWAY_MAX_JNR":        "PxJoin",
    "ADJUST_PRTY_TFM":     "CTransformerStage",
    "TE_EMPLOYEE_PHONE_DW":"TeradataConnectorPX",
    "NULL_VAL_TFM":        "CTransformerStage",
    "WSSEN_ERR_SEQ":       "PxSequentialFile",
}

# Links derived from XML lines 144 (LinkNames) and 149 (TargetStageIDs) - EXACT mapping verified
# Format: (source_stage, target_stage, link_name)
LINKS = [
    # WSTELE_LND_TDC
    ("WSTELE_LND_TDC",      "SEPARATE_PHTYPE_TFM",  "LNK_OUT_TDC"),
    # SEPARATE_PHTYPE_TFM -> 6 outputs
    ("SEPARATE_PHTYPE_TFM", "ALL_REC_FNL",           "EMERGENCY_IN_TFM"),
    ("SEPARATE_PHTYPE_TFM", "ALL_REC_FNL",           "TEMP_TYPE_TFM"),
    ("SEPARATE_PHTYPE_TFM", "BASIC_TYPE_PVT",        "BASIC_TYPE_TFM"),
    ("SEPARATE_PHTYPE_TFM", "HOME_CPY",              "HOME_TYPE_CPY"),
    ("SEPARATE_PHTYPE_TFM", "AWAY_CPY",              "AWAY_TYPE_CPY"),
    ("SEPARATE_PHTYPE_TFM", "WSSEN_ERR_SEQ",         "LNK_OUT_ERR_SEQ"),
    # BASIC_TYPE_PVT
    ("BASIC_TYPE_PVT",      "BASIC_REC_TFM",         "BREC_OUT_PVT"),
    # BASIC_REC_TFM -> 5 outputs (ALL_REC_FNL + right-inputs to 4 joins)
    ("BASIC_REC_TFM",       "ALL_REC_FNL",           "ALL_BREC_OUT_TFM"),
    ("BASIC_REC_TFM",       "HOME_BASIC_JNR",        "CALL_PRIORITY_RIGHT_JNR"),
    ("BASIC_REC_TFM",       "BASIC_MAX_JNR",         "BREC_OUT_TFM"),
    ("BASIC_REC_TFM",       "AWAY_BASIC_JNR",        "CALL_PRI_RIGHT_JNR"),
    ("BASIC_REC_TFM",       "AWAY_MAX_JNR",          "BASIC_REC_OUT_TFM"),
    # HOME branch: HOME_SEQ1_PVT output -> HOME_SEQ_FNL (verified: V0S35->HSEQ1_OUT_PVT->V0S46=HOME_SEQ_FNL)
    ("HOME_SEQ1_PVT",       "HOME_SEQ_FNL",          "HSEQ1_OUT_PVT"),
    # HOME_CPY -> 3 pivot branches
    ("HOME_CPY",            "HOME_SEQ1_PVT",         "HOME_SEQ1_CPY"),
    ("HOME_CPY",            "HOME_SEQ2_PVT",         "HOME_SEQ2_CPY"),
    ("HOME_CPY",            "HOME_SEQ3_PVT",         "HOME_SEQ3_CPY"),
    # HOME_SEQ2_PVT -> HOME_SEQ2_TFM (verified: V4S2->HSEQ2_OUT_PVT->V0S72=HOME_SEQ2_TFM)
    ("HOME_SEQ2_PVT",       "HOME_SEQ2_TFM",         "HSEQ2_OUT_PVT"),
    # HOME_SEQ3_PVT -> HOME_SEQ3_TFM (verified: V5S1->HSEQ3_OUT_PVT->V0S73=HOME_SEQ3_TFM)
    ("HOME_SEQ3_PVT",       "HOME_SEQ3_TFM",         "HSEQ3_OUT_PVT"),
    # HOME_SEQ2_TFM, HOME_SEQ3_TFM -> HOME_SEQ_FNL (verified: V0S72->V0S46, V0S73->V0S46)
    ("HOME_SEQ2_TFM",       "HOME_SEQ_FNL",          "HSEQ2_OUT_TFM"),
    ("HOME_SEQ3_TFM",       "HOME_SEQ_FNL",          "HSEQ3_OUT_TFM"),
    # HOME_SEQ_FNL -> INC_PRTY_TFM (verified: V0S46->HSEQ_OUT_FNL->V0S51=INC_PRTY_TFM)
    ("HOME_SEQ_FNL",        "INC_PRTY_TFM",          "HSEQ_OUT_FNL"),
    # INC_PRTY_TFM -> HOME_BASIC_JNR (verified: V0S51->CALL_PRIORITY_LEFT_JNR->V0S60=HOME_BASIC_JNR)
    ("INC_PRTY_TFM",        "HOME_BASIC_JNR",        "CALL_PRIORITY_LEFT_JNR"),
    # HOME_BASIC_JNR -> NEW_PRTY_TFM (verified: V0S60->HBREC_OUT_JNR->V0S65=NEW_PRTY_TFM)
    ("HOME_BASIC_JNR",      "NEW_PRTY_TFM",          "HBREC_OUT_JNR"),
    # NEW_PRTY_TFM -> ALL_REC_FNL + MAX_PRTY_AGG (verified: V0S65->NPRTY_OUT_TFM->V0S106, CPRTY_OUT_TFM->V0S77)
    ("NEW_PRTY_TFM",        "ALL_REC_FNL",           "NPRTY_OUT_TFM"),
    ("NEW_PRTY_TFM",        "MAX_PRTY_AGG",          "CPRTY_OUT_TFM"),
    # MAX_PRTY_AGG -> BASIC_MAX_JNR
    ("MAX_PRTY_AGG",        "BASIC_MAX_JNR",         "HMAX_OUT_AGG"),
    # BASIC_MAX_JNR -> ADJUSTING_PRTY_TFM
    ("BASIC_MAX_JNR",       "ADJUSTING_PRTY_TFM",    "BMAX_OUT_JNR"),
    # ADJUSTING_PRTY_TFM -> ALL_REC_FNL
    ("ADJUSTING_PRTY_TFM",  "ALL_REC_FNL",           "ADJ_PRTY_OUT_TFM"),
    # AWAY branch (verified from TargetStageIDs)
    # AWAY_CPY -> 3 pivot branches
    ("AWAY_CPY",            "AWAY_SEQ1_PVT",         "AWAY_SEQ1_CPY"),
    ("AWAY_CPY",            "AWAY_SEQ2_PVT",         "AWAY_SEQ2_CPY"),
    ("AWAY_CPY",            "AWAY_SEQ3_PVT",         "AWAY_SEQ3_CPY"),
    # AWAY_SEQ1_PVT -> AWAY_SEQ_FNL (verified: V6S8->ASEQ1_OUT_PVT->V6S2=AWAY_SEQ_FNL)
    ("AWAY_SEQ1_PVT",       "AWAY_SEQ_FNL",          "ASEQ1_OUT_PVT"),
    # AWAY_SEQ2_PVT -> AWAY_SEQ2_TFM (verified: V6S5->ASEQ2_OUT_PVT->V6S1=AWAY_SEQ2_TFM)
    ("AWAY_SEQ2_PVT",       "AWAY_SEQ2_TFM",         "ASEQ2_OUT_PVT"),
    # AWAY_SEQ3_PVT -> AWAY_SEQ3_TFM (verified: V6S4->ASEQ3_OUT_PVT->V6S0=AWAY_SEQ3_TFM)
    ("AWAY_SEQ3_PVT",       "AWAY_SEQ3_TFM",         "ASEQ3_OUT_PVT"),
    # AWAY_SEQ2_TFM, AWAY_SEQ3_TFM -> AWAY_SEQ_FNL (verified: V6S1->V6S2, V6S0->V6S2)
    ("AWAY_SEQ2_TFM",       "AWAY_SEQ_FNL",          "ASEQ2_OUT_TFM"),
    ("AWAY_SEQ3_TFM",       "AWAY_SEQ_FNL",          "ASEQ3_OUT_TFM"),
    # AWAY_SEQ_FNL -> INC_AWAY_PRTY_TFM (verified: V6S2->ASEQ_OUT_FNL->V7S0)
    ("AWAY_SEQ_FNL",        "INC_AWAY_PRTY_TFM",     "ASEQ_OUT_FNL"),
    # INC_AWAY_PRTY_TFM -> AWAY_BASIC_JNR (verified: V7S0->CALL_PRIORITY_LEFT_JNR->V8S0)
    ("INC_AWAY_PRTY_TFM",   "AWAY_BASIC_JNR",        "CALL_PRIORITY_LEFT_JNR"),
    # AWAY_BASIC_JNR -> ABREC_NEW_PRTY_TFM (verified: V8S0->ABREC_OUT_JNR->V9S2)
    ("AWAY_BASIC_JNR",      "ABREC_NEW_PRTY_TFM",    "ABREC_OUT_JNR"),
    # MAX_ABPRTY_AGG -> AWAY_MAX_JNR (verified: V9S0->AMAX_OUT_AGG->V10S0)
    ("MAX_ABPRTY_AGG",      "AWAY_MAX_JNR",          "AMAX_OUT_AGG"),
    # ABREC_NEW_PRTY_TFM -> ALL_REC_FNL + MAX_ABPRTY_AGG (verified: V9S2->ABNEW_PRTY_OUT_TFM->V0S106, ABCPRTY_OUT_TFM->V9S0)
    ("ABREC_NEW_PRTY_TFM",  "ALL_REC_FNL",           "ABNEW_PRTY_OUT_TFM"),
    ("ABREC_NEW_PRTY_TFM",  "MAX_ABPRTY_AGG",        "ABCPRTY_OUT_TFM"),
    # AWAY_MAX_JNR -> ADJUST_PRTY_TFM (verified: V10S0->DSLink85->V11S0)
    ("AWAY_MAX_JNR",        "ADJUST_PRTY_TFM",       "DSLink85"),
    # ADJUST_PRTY_TFM -> ALL_REC_FNL (verified: V11S0->ADJUST_PRTY_OUT_TFM->V0S106)
    ("ADJUST_PRTY_TFM",     "ALL_REC_FNL",           "ADJUST_PRTY_OUT_TFM"),
    # ALL_REC_FNL -> NULL_VAL_TFM
    ("ALL_REC_FNL",         "NULL_VAL_TFM",          "ALL_REC_OUT_FNL"),
    # NULL_VAL_TFM -> TE_EMPLOYEE_PHONE_DW
    ("NULL_VAL_TFM",        "TE_EMPLOYEE_PHONE_DW",  "LNK_IN_DW"),
]

# ---------------------------------------------------------------------------
# Build adjacency structures
# ---------------------------------------------------------------------------
def build_graph(links):
    out_links = defaultdict(list)   # stage -> [(target, link_name)]
    in_links  = defaultdict(list)   # stage -> [(source, link_name)]
    for src, tgt, lnk in links:
        out_links[src].append((tgt, lnk))
        in_links[tgt].append((src, lnk))
    return out_links, in_links


def assign_sequence_numbers(stage_names, out_links, in_links):
    """Topological sort with parallel branch labelling."""
    # Kahn's algorithm
    in_degree = {s: len(in_links[s]) for s in stage_names}
    queue = deque([s for s in stage_names if in_degree[s] == 0])
    order = []
    while queue:
        # All nodes at same depth are in queue simultaneously -> form a "level"
        level = list(queue)
        queue.clear()
        order.append(level)
        for node in level:
            for (tgt, _) in out_links[node]:
                in_degree[tgt] -= 1
                if in_degree[tgt] == 0:
                    queue.append(tgt)

    seq_map = {}
    seq_num = 1
    for level in order:
        if len(level) == 1:
            seq_map[level[0]] = str(seq_num)
        else:
            for i, stage in enumerate(level):
                seq_map[stage] = f"{seq_num}{chr(65 + i)}"
        seq_num += 1

    return seq_map


def get_category(name, stage_type, in_links, out_links):
    has_in  = len(in_links[name]) > 0
    has_out = len(out_links[name]) > 0
    if not has_in:
        return "Source"
    if not has_out:
        return "Target"
    return "Transformation"


def get_transformation_summary(name, stage_type, in_links, out_links):
    n_in  = len(in_links[name])
    n_out = len(out_links[name])

    summaries = {
        "WSTELE_LND_TDC":      "Read from Teradata table CREW_WSTELE_LND",
        "TE_EMPLOYEE_PHONE_DW":"Write to Teradata table TE_EMPLOYEE_PHONE",
        "WSSEN_ERR_SEQ":       "Write rejected records to sequential error file",
        "SEPARATE_PHTYPE_TFM": "Column derivation / conditional routing — routes records by phone type (EMERGENCY, TEMP, BASIC, HOME, AWAY, ERR) with EMP_NBR validity check",
        "BASIC_REC_TFM":       "Column derivation — maps BASIC phone records to standard columns (EMP_NBR, PH_NBR, PH_TYPE, HOME_AWAY_IND, UNLISTD_IND, CALL_PRTY, etc.)",
        "INC_PRTY_TFM":        "Column derivation — increments CALL_PRIORITY by pivot index offset (+1) to generate 1-based priorities",
        "NEW_PRTY_TFM":        "Column derivation — assigns new call list priorities to home phone sequence records",
        "HOME_SEQ2_TFM":       "Column derivation — transforms SEQ2 home records, adjusts priority offset (+5)",
        "HOME_SEQ3_TFM":       "Column derivation — transforms SEQ3 home records, adjusts priority offset (+10)",
        "ADJUSTING_PRTY_TFM":  "Column derivation — calculates adjusted priority (PH_PRTY) for basic records based on max priority join",
        "INC_AWAY_PRTY_TFM":   "Column derivation — increments AWAY call priority by pivot index offset",
        "ABREC_NEW_PRTY_TFM":  "Column derivation — assigns new call priorities to AWAY phone sequence records",
        "AWAY_SEQ2_TFM":       "Column derivation — transforms AWAY SEQ2 records, adjusts priority offset (+5)",
        "AWAY_SEQ3_TFM":       "Column derivation — transforms AWAY SEQ3 records, adjusts priority offset (+10)",
        "ADJUST_PRTY_TFM":     "Column derivation — calculates adjusted priority for AWAY basic records based on max priority join",
        "NULL_VAL_TFM":        "Column derivation / null-handling — applies final null checks, pads EMP_NO to 9 chars, formats EFF_START_TM/EFF_END_TM to HH:MM:SS",
        "BASIC_TYPE_PVT":      "Pivot 5 BASIC phone columns per employee into individual rows (BASIC_PH_NBR_1..5)",
        "HOME_SEQ1_PVT":       "Pivot 5 HOME sequence-1 priority columns per employee into individual rows",
        "HOME_SEQ2_PVT":       "Pivot 5 HOME sequence-2 priority columns per employee into individual rows",
        "HOME_SEQ3_PVT":       "Pivot 5 HOME sequence-3 priority columns per employee into individual rows",
        "AWAY_SEQ1_PVT":       "Pivot 5 AWAY sequence-1 priority columns per employee into individual rows",
        "AWAY_SEQ2_PVT":       "Pivot 5 AWAY sequence-2 priority columns per employee into individual rows",
        "AWAY_SEQ3_PVT":       "Pivot 5 AWAY sequence-3 priority columns per employee into individual rows",
        "HOME_CPY":            "Copy HOME stream to 3 outputs (SEQ1, SEQ2, SEQ3 pivot branches)",
        "AWAY_CPY":            "Copy AWAY stream to 3 outputs (SEQ1, SEQ2, SEQ3 pivot branches)",
        "HOME_SEQ_FNL":        "Merge 3 HOME pivot streams (SEQ1+SEQ2+SEQ3) into single stream",
        "AWAY_SEQ_FNL":        "Merge 3 AWAY pivot streams (SEQ1+SEQ2+SEQ3) into single stream",
        "ALL_REC_FNL":         "Merge all phone type streams (EMERGENCY, TEMP, BASIC-normal, HOME-normal, AWAY-normal, HOME-adjusted, AWAY-adjusted) into single stream",
        "HOME_BASIC_JNR":      "Join: left-input=HOME pivot stream (INC_PRTY_TFM), right-input=BASIC records (BASIC_REC_TFM) on EMP_NBR. Output: HBREC_OUT_JNR → NEW_PRTY_TFM",
        "BASIC_MAX_JNR":       "Join: left-input=BASIC detail (BREC_OUT_TFM from BASIC_REC_TFM), right-input=MAX priority (HMAX_OUT_AGG from MAX_PRTY_AGG) on EMP_NBR",
        "AWAY_BASIC_JNR":      "Join: left-input=AWAY pivot stream (INC_AWAY_PRTY_TFM), right-input=BASIC records (BASIC_REC_TFM) on EMP_NBR. Output: ABREC_OUT_JNR → ABREC_NEW_PRTY_TFM",
        "AWAY_MAX_JNR":        "Join: left-input=AWAY detail (BASIC_REC_OUT_TFM from BASIC_REC_TFM), right-input=MAX AWAY priority (AMAX_OUT_AGG from MAX_ABPRTY_AGG) on EMP_NBR",
        "MAX_PRTY_AGG":        "Aggregate: MAX(CALL_PRTY) GROUP BY EMP_NBR for HOME records",
        "MAX_ABPRTY_AGG":      "Aggregate: MAX(CALL_PRTY) GROUP BY EMP_NBR for AWAY records",
    }
    return summaries.get(name, f"{stage_type} transformation")


# ---------------------------------------------------------------------------
# Parse XML to get column-level derivations
# ---------------------------------------------------------------------------
def parse_xml_derivations(xml_path):
    """
    Returns a dict: stage_name -> {col_name: derivation_expr}
    Also returns filter_constraints: stage_name -> constraint_expr
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()

    derivations     = {}  # stage -> {col: expr}
    constraints     = {}  # stage -> constraint list
    stage_id_to_name = {}

    # Build id->name map from Record elements
    for record in root.iter("Record"):
        ident = record.get("Identifier", "")
        name_prop = record.find("./Property[@Name='Name']")
        if name_prop is not None and name_prop.text:
            stage_id_to_name[ident] = name_prop.text

    # Now find TrxOutput records to get derivations
    for record in root.iter("Record"):
        rec_type = record.get("Type", "")
        if rec_type not in ("TrxOutput", "CustomOutput"):
            continue

        link_name_prop = record.find("./Property[@Name='Name']")
        link_name = link_name_prop.text if link_name_prop is not None else ""

        # Constraint
        constraint_prop = record.find("./Property[@Name='Constraint']")
        if constraint_prop is not None and constraint_prop.text:
            constraints[link_name] = constraint_prop.text

        # Columns
        for col in record.iter("SubRecord"):
            col_name_prop = col.find("Property[@Name='Name']")
            deriv_prop    = col.find("Property[@Name='Derivation']")
            if col_name_prop is not None and deriv_prop is not None:
                col_name = col_name_prop.text or ""
                deriv    = deriv_prop.text or ""
                if col_name and deriv:
                    if link_name not in derivations:
                        derivations[link_name] = {}
                    derivations[link_name][col_name] = deriv

    return derivations, constraints, stage_id_to_name


# ---------------------------------------------------------------------------
# Source columns from the SQL SELECT in XML
# ---------------------------------------------------------------------------
SOURCE_COLUMNS = [
    "EMP_NBR", "USER_ID", "TELE_LAST_UPDATED_DATE", "TELE_LAST_UPDATED_TIME",
    "PH_COMMENTS",
    "PH_NBR",                  # concatenated EMERGENCY phone
    "TEMP_PH_NBR",             # TEMP phone number
    "TEMP_PH_ACCESS", "TEMP_PH_COMMENTS", "TELE_TEMP_PH_DATE", "TELE_TEMP_PH_TIME",
    # BASIC 1-5
    "BASIC_PH_NBR_1","BASIC_PH_ACCESS_1","BASIC_PH_COMMENTS_1","BASIC_PH_TYPE_1","BASIC_PH_UNLIST_CD_1","BASIC_PH_HOME_AWAY_CD_1",
    "BASIC_PH_NBR_2","BASIC_PH_ACCESS_2","BASIC_PH_COMMENTS_2","BASIC_PH_TYPE_2","BASIC_PH_UNLIST_CD_2","BASIC_PH_HOME_AWAY_CD_2",
    "BASIC_PH_NBR_3","BASIC_PH_ACCESS_3","BASIC_PH_COMMENTS_3","BASIC_PH_TYPE_3","BASIC_PH_UNLIST_CD_3","BASIC_PH_HOME_AWAY_CD_3",
    "BASIC_PH_NBR_4","BASIC_PH_ACCESS_4","BASIC_PH_COMMENTS_4","BASIC_PH_TYPE_4","BASIC_PH_UNLIST_CD_4","BASIC_PH_HOME_AWAY_CD_4",
    "BASIC_PH_NBR_5","BASIC_PH_ACCESS_5","BASIC_PH_COMMENTS_5","BASIC_PH_TYPE_5","BASIC_PH_UNLIST_CD_5","BASIC_PH_HOME_AWAY_CD_5",
    # HOME 1-3 sets
    "TELE_HOME_PRI_FROM_1","TELE_HOME_PRI_TO_1","TELE_HOME_PRI_SEQ_1_1","TELE_HOME_PRI_SEQ_2_1","TELE_HOME_PRI_SEQ_3_1","TELE_HOME_PRI_SEQ_4_1","TELE_HOME_PRI_SEQ_5_1",
    "TELE_HOME_PRI_FROM_2","TELE_HOME_PRI_TO_2","TELE_HOME_PRI_SEQ_1_2","TELE_HOME_PRI_SEQ_2_2","TELE_HOME_PRI_SEQ_3_2","TELE_HOME_PRI_SEQ_4_2","TELE_HOME_PRI_SEQ_5_2",
    "TELE_HOME_PRI_FROM_3","TELE_HOME_PRI_TO_3","TELE_HOME_PRI_SEQ_1_3","TELE_HOME_PRI_SEQ_2_3","TELE_HOME_PRI_SEQ_3_3","TELE_HOME_PRI_SEQ_4_3","TELE_HOME_PRI_SEQ_5_3",
    # AWAY 1-3 sets
    "TELE_AWAY_PRI_FROM_1","TELE_AWAY_PRI_TO_1","TELE_AWAY_PRI_SEQ_1_1","TELE_AWAY_PRI_SEQ_2_1","TELE_AWAY_PRI_SEQ_3_1","TELE_AWAY_PRI_SEQ_4_1","TELE_AWAY_PRI_SEQ_5_1",
    "TELE_AWAY_PRI_FROM_2","TELE_AWAY_PRI_TO_2","TELE_AWAY_PRI_SEQ_1_2","TELE_AWAY_PRI_SEQ_2_2","TELE_AWAY_PRI_SEQ_3_2","TELE_AWAY_PRI_SEQ_4_2","TELE_AWAY_PRI_SEQ_5_2",
    "TELE_AWAY_PRI_FROM_3","TELE_AWAY_PRI_TO_3","TELE_AWAY_PRI_SEQ_1_3","TELE_AWAY_PRI_SEQ_2_3","TELE_AWAY_PRI_SEQ_3_3","TELE_AWAY_PRI_SEQ_4_3","TELE_AWAY_PRI_SEQ_5_3",
]

# Target columns with their derivations (from NULL_VAL_TFM -> TE_EMPLOYEE_PHONE_DW, link LNK_IN_DW)
TARGET_COLUMNS_MAP = {
    "EMP_NBR":       {
        "derivation": "ALL_REC_OUT_FNL.EMP_NBR",
        "source_col": "EMP_NBR",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "EMP_NO":        {
        "derivation": "If Len(TrimLeadingTrailing(ALL_REC_OUT_FNL.EMP_NBR))=9 Then TrimLeadingTrailing(ALL_REC_OUT_FNL.EMP_NBR) Else Str('0',9-Len(TrimLeadingTrailing(ALL_REC_OUT_FNL.EMP_NBR))):TrimLeadingTrailing(ALL_REC_OUT_FNL.EMP_NBR)",
        "source_col": "EMP_NBR",
        "mapping_type": "Derived",
        "trfm_types": "String padding / conditional",
    },
    "PH_LIST":       {
        "derivation": "ALL_REC_OUT_FNL.PH_LIST",
        "source_col": "PH_LIST (CALL_LIST from SEPARATE_PHTYPE_TFM)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "PH_PRTY":       {
        "derivation": "ALL_REC_OUT_FNL.PH_PRTY",
        "source_col": "PH_PRTY (CALL_PRTY derived through priority chain)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "EFF_START_TM":  {
        "derivation": "left(ALL_REC_OUT_FNL.EFF_START_TM,2):':':right(ALL_REC_OUT_FNL.EFF_START_TM,2):':00'",
        "source_col": "EFF_START_TM (constant '0000' from SEPARATE_PHTYPE_TFM)",
        "mapping_type": "Derived",
        "trfm_types": "String formatting",
    },
    "EFF_END_TM":    {
        "derivation": "left(ALL_REC_OUT_FNL.EFF_END_TM,2):':':right(ALL_REC_OUT_FNL.EFF_END_TM,2):':00'",
        "source_col": "EFF_END_TM (constant '2359' from SEPARATE_PHTYPE_TFM)",
        "mapping_type": "Derived",
        "trfm_types": "String formatting",
    },
    "PH_NBR":        {
        "derivation": "ALL_REC_OUT_FNL.PH_NBR",
        "source_col": "PH_NBR (concatenated from source: area+prefix+number)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "PH_ACCESS":     {
        "derivation": "if TrimLeadingTrailing(nulltovalue(ALL_REC_OUT_FNL.PH_ACCESS,''))='' or TrimLeadingTrailing(nulltovalue(ALL_REC_OUT_FNL.PH_ACCESS,'')) < ' ' then setnull() else TrimLeadingTrailing(nulltovalue(ALL_REC_OUT_FNL.PH_ACCESS,''))",
        "source_col": "PH_ACCESS",
        "mapping_type": "Derived",
        "trfm_types": "Null handling / trim",
    },
    "PH_COMMENTS":   {
        "derivation": "if TrimLeadingTrailing(nulltovalue(ALL_REC_OUT_FNL.PH_COMMENTS,''))='' or TrimLeadingTrailing(nulltovalue(ALL_REC_OUT_FNL.PH_COMMENTS,'')) < ' ' then setnull() else TrimLeadingTrailing(nulltovalue(ALL_REC_OUT_FNL.PH_COMMENTS,''))",
        "source_col": "PH_COMMENTS",
        "mapping_type": "Derived",
        "trfm_types": "Null handling / trim",
    },
    "PH_TYPE":       {
        "derivation": "ALL_REC_OUT_FNL.PH_TYPE",
        "source_col": "PH_TYPE (BASIC_PH_TYPE_n from pivot, null for EMERGENCY/TEMP)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "UNLISTD_IND":   {
        "derivation": "ALL_REC_OUT_FNL.UNLISTD_IND",
        "source_col": "UNLISTD_IND (BASIC_PH_UNLIST_CD_n from pivot, null for EMERGENCY/TEMP)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "HOME_AWAY_IND": {
        "derivation": "ALL_REC_OUT_FNL.HOME_AWAY_IND",
        "source_col": "HOME_AWAY_IND (BASIC_PH_HOME_AWAY_CD_n, null for EMERGENCY/TEMP)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "TEMP_PH_EXP_TS":{
        "derivation": "ALL_REC_OUT_FNL.TEMP_PH_EXP_TS",
        "source_col": "TEMP_PH_EXP_TS (timestamp from TELE_TEMP_PH_DATE/TIME)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "TD_LD_TS":      {
        "derivation": "CurrentTimestamp()",
        "source_col": "N/A (system function)",
        "mapping_type": "Constant",
        "trfm_types": "System timestamp",
    },
    "USER_ID":       {
        "derivation": "ALL_REC_OUT_FNL.USER_ID",
        "source_col": "USER_ID (TELE_LAST_UPDATED_BY from source)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
    "UPD_TS":        {
        "derivation": "ALL_REC_OUT_FNL.UPD_TS",
        "source_col": "UPD_TS (timestamp derived from TELE_LAST_UPDATED_DATE+TIME)",
        "mapping_type": "Pass-through",
        "trfm_types": "Pass-through",
    },
}

# Error stage columns
ERROR_COLUMNS = [
    "EMP_NBR", "PH_NBR", "TEMP_PH_NBR",
    "BASIC_PH_NBR_1","BASIC_PH_NBR_2","BASIC_PH_NBR_3","BASIC_PH_NBR_4","BASIC_PH_NBR_5",
]

# ---------------------------------------------------------------------------
# Stage path description helpers
# ---------------------------------------------------------------------------
STAGE_PATH_MAP = {
    "EMP_NBR":       "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → [various branches] → ALL_REC_FNL → NULL_VAL_TFM → TE_EMPLOYEE_PHONE_DW",
    "EMP_NO":        "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → [various branches] → ALL_REC_FNL → NULL_VAL_TFM → TE_EMPLOYEE_PHONE_DW",
    "PH_LIST":       "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM (assigns CALL_LIST constant) → [branches] → ALL_REC_FNL → NULL_VAL_TFM → TE_EMPLOYEE_PHONE_DW",
    "PH_PRTY":       "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → BASIC_TYPE_PVT → BASIC_REC_TFM → [pivot/join/agg chain] → ADJUSTING_PRTY_TFM → ALL_REC_FNL → NULL_VAL_TFM",
    "EFF_START_TM":  "SEPARATE_PHTYPE_TFM (constant '0000') → ALL_REC_FNL → NULL_VAL_TFM → TE_EMPLOYEE_PHONE_DW",
    "EFF_END_TM":    "SEPARATE_PHTYPE_TFM (constant '2359') → ALL_REC_FNL → NULL_VAL_TFM → TE_EMPLOYEE_PHONE_DW",
    "PH_NBR":        "WSTELE_LND_TDC (concatenated area+prefix+num per type) → SEPARATE_PHTYPE_TFM → ALL_REC_FNL → NULL_VAL_TFM",
    "PH_ACCESS":     "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → BASIC_REC_TFM (BASIC_PH_ACCESS_n) → ALL_REC_FNL → NULL_VAL_TFM",
    "PH_COMMENTS":   "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → ALL_REC_FNL → NULL_VAL_TFM",
    "PH_TYPE":       "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → BASIC_REC_TFM (BASIC_PH_TYPE_n) → ALL_REC_FNL → NULL_VAL_TFM",
    "UNLISTD_IND":   "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → BASIC_REC_TFM (BASIC_PH_UNLIST_CD_n) → ALL_REC_FNL → NULL_VAL_TFM",
    "HOME_AWAY_IND": "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → BASIC_REC_TFM (BASIC_PH_HOME_AWAY_CD_n) → ALL_REC_FNL → NULL_VAL_TFM",
    "TEMP_PH_EXP_TS":"WSTELE_LND_TDC (TELE_TEMP_PH_DATE, TELE_TEMP_PH_TIME) → SEPARATE_PHTYPE_TFM → ALL_REC_FNL → NULL_VAL_TFM",
    "TD_LD_TS":      "N/A — system constant CurrentTimestamp() in NULL_VAL_TFM",
    "USER_ID":       "WSTELE_LND_TDC (TELE_LAST_UPDATED_BY) → SEPARATE_PHTYPE_TFM → ALL_REC_FNL → NULL_VAL_TFM",
    "UPD_TS":        "WSTELE_LND_TDC (TELE_LAST_UPDATED_DATE, TELE_LAST_UPDATED_TIME) → SEPARATE_PHTYPE_TFM → ALL_REC_FNL → NULL_VAL_TFM",
}

JOIN_LOGIC_MAP = {
    "PH_PRTY": "HOME_BASIC_JNR: LEFT JOIN HOME pivot records WITH BASIC records ON EMP_NBR; BASIC_MAX_JNR: INNER JOIN ON EMP_NBR with MAX aggregated priority; AWAY_BASIC_JNR: LEFT JOIN AWAY pivot records WITH BASIC records ON EMP_NBR; AWAY_MAX_JNR: INNER JOIN ON EMP_NBR",
}

FILTER_LOGIC_MAP = {
    "EMP_NBR":  "SEPARATE_PHTYPE_TFM constraint: is_valid('INT32', EMP_NBR) AND is_valid('int64', PH_NBR) [EMERGENCY]; error branch captures invalid records",
    "PH_NBR":   "NULL_VAL_TFM output constraint: isvalid('int64', ALL_REC_OUT_FNL.PH_NBR) — records with invalid PH_NBR are routed to WSSEN_ERR_SEQ",
}

AGG_LOGIC_MAP = {
    "PH_PRTY": "MAX_PRTY_AGG: MAX(CALL_PRTY) GROUP BY EMP_NBR for HOME records; MAX_ABPRTY_AGG: MAX(CALL_PRTY) GROUP BY EMP_NBR for AWAY records",
}

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
HDR_FILL  = PatternFill("solid", fgColor="4472C4")
HDR_FONT  = Font(bold=True, color="FFFFFF")
ALT_FILL  = PatternFill("solid", fgColor="F5F5F5")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
HDR_ALIGN  = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_sheet(ws, col_widths):
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if row_idx == 1:
                cell.font  = HDR_FONT
                cell.fill  = HDR_FILL
                cell.alignment = HDR_ALIGN
            else:
                cell.alignment = WRAP_ALIGN
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

    ws.row_dimensions[1].height = 30


# ---------------------------------------------------------------------------
# Build Mermaid diagram
# ---------------------------------------------------------------------------
def build_mermaid(stage_names, links, out_links, in_links):
    """Return the Mermaid flowchart string."""
    lines = ["flowchart LR", ""]

    # Node declarations
    lines.append("    %% Source stages")
    for name in stage_names:
        stype = STAGE_TYPE_IDS.get(name, "Unknown")
        cat   = get_category(name, stype, in_links, out_links)
        label = f"{name}\\n({stype})"
        if cat == "Source":
            lines.append(f'    {name}(["{label}"])')
        elif cat == "Target":
            lines.append(f'    {name}[("{label}")]')
        else:
            lines.append(f'    {name}["{label}"]')

    lines.append("")
    lines.append("    %% Edges")

    # Deduplicate edges for Mermaid (keep unique src->tgt->link combos)
    seen = set()
    for src, tgt, lnk in links:
        key = (src, tgt, lnk)
        if key in seen:
            continue
        seen.add(key)
        lines.append(f'    {src} -->|"{lnk}"| {tgt}')

    lines.append("")
    lines.append("    %% Style")
    lines.append('    classDef source fill:#d4edda,stroke:#28a745,color:#000,font-weight:bold')
    lines.append('    classDef target fill:#f8d7da,stroke:#dc3545,color:#000,font-weight:bold')
    lines.append('    classDef xform fill:#cce5ff,stroke:#004085,color:#000')
    lines.append("")

    sources = [n for n in stage_names if get_category(n, STAGE_TYPE_IDS.get(n), in_links, out_links) == "Source"]
    targets = [n for n in stage_names if get_category(n, STAGE_TYPE_IDS.get(n), in_links, out_links) == "Target"]
    xforms  = [n for n in stage_names if get_category(n, STAGE_TYPE_IDS.get(n), in_links, out_links) == "Transformation"]

    if sources:
        lines.append(f'    class {",".join(sources)} source')
    if targets:
        lines.append(f'    class {",".join(targets)} target')
    if xforms:
        lines.append(f'    class {",".join(xforms)} xform')

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Lineage summary (plain English)
# ---------------------------------------------------------------------------
LINEAGE_SUMMARY = """Plain-English Lineage Summary
==============================

1. WSTELE_LND_TDC (Source)
   Reads from Teradata table #ENV.$TD_WORK_DB#.CREW_WSTELE_LND via a SQL SELECT.
   The SELECT pre-concatenates phone number components (area+prefix+number) and
   aliases columns for EMERGENCY, TEMP, BASIC (x5), HOME (3 sets x 5 seq), and
   AWAY (3 sets x 5 seq) phone types.

2. SEPARATE_PHTYPE_TFM (CTransformerStage)
   Routes each source row to up to 6 output streams based on phone type and
   EMP_NBR/PH_NBR validity:
   - EMERGENCY_IN_TFM  → ALL_REC_FNL  (if EMP_NBR int-valid AND emergency PH_NBR int-valid)
   - TEMP_TYPE_TFM     → ALL_REC_FNL  (if EMP_NBR int-valid AND TEMP_PH_NBR valid)
   - BASIC_TYPE_TFM    → BASIC_TYPE_PVT
   - HOME_TYPE_CPY     → HOME_CPY
   - AWAY_TYPE_CPY     → AWAY_CPY
   - LNK_OUT_ERR_SEQ   → WSSEN_ERR_SEQ (records failing all validity checks)

3. BASIC branch (BASIC_TYPE_PVT → BASIC_REC_TFM)
   BASIC_TYPE_PVT pivots 5 BASIC phone columns per employee into individual rows.
   BASIC_REC_TFM maps each pivoted row to standard columns (EMP_NBR, PH_NBR,
   PH_TYPE, HOME_AWAY_IND, UNLISTD_IND) and sets CALL_PRTY=1 (initial).
   Output feeds: ALL_REC_FNL, HOME_BASIC_JNR (right), BASIC_MAX_JNR (detail),
   AWAY_BASIC_JNR (right), AWAY_MAX_JNR (detail).

4. HOME priority chain
   HOME_CPY → 3 copies → HOME_SEQ1_PVT / HOME_SEQ2_PVT / HOME_SEQ3_PVT
   Each pivot expands 5 HOME sequence priority columns per employee per set.
   HOME_SEQ1_PVT output (HSEQ1_OUT_PVT) feeds directly into HOME_SEQ_FNL.
   HOME_SEQ2_PVT → HOME_SEQ2_TFM (priority offset +5) → HOME_SEQ_FNL.
   HOME_SEQ3_PVT → HOME_SEQ3_TFM (priority offset +10) → HOME_SEQ_FNL.
   HOME_SEQ_FNL (HSEQ_OUT_FNL) → INC_PRTY_TFM: increments CALL_PRIORITY (1-based).
   HOME_BASIC_JNR: left=INC_PRTY_TFM, right=BASIC_REC_TFM on EMP_NBR.
   NEW_PRTY_TFM: NPRTY_OUT_TFM → ALL_REC_FNL; CPRTY_OUT_TFM → MAX_PRTY_AGG.
   MAX_PRTY_AGG: MAX(CALL_PRTY) GROUP BY EMP_NBR → BASIC_MAX_JNR.
   BASIC_MAX_JNR: left=BASIC_REC_TFM(BREC_OUT_TFM), right=MAX_PRTY_AGG on EMP_NBR.
   ADJUSTING_PRTY_TFM: Final priority calculation → ALL_REC_FNL.

5. AWAY priority chain (mirrors HOME structure)
   AWAY_CPY → 3 copies → AWAY_SEQ1_PVT / AWAY_SEQ2_PVT / AWAY_SEQ3_PVT
   AWAY_SEQ1_PVT → AWAY_SEQ_FNL directly.
   AWAY_SEQ2_PVT → AWAY_SEQ2_TFM (priority offset +5) → AWAY_SEQ_FNL.
   AWAY_SEQ3_PVT → AWAY_SEQ3_TFM (priority offset +10) → AWAY_SEQ_FNL.
   AWAY_SEQ_FNL (ASEQ_OUT_FNL) → INC_AWAY_PRTY_TFM: increments AWAY CALL_PRIORITY.
   AWAY_BASIC_JNR: left=INC_AWAY_PRTY_TFM, right=BASIC_REC_TFM on EMP_NBR.
   ABREC_NEW_PRTY_TFM: ABNEW_PRTY_OUT_TFM → ALL_REC_FNL; ABCPRTY_OUT_TFM → MAX_ABPRTY_AGG.
   MAX_ABPRTY_AGG: MAX(CALL_PRTY) GROUP BY EMP_NBR → AWAY_MAX_JNR.
   AWAY_MAX_JNR: left=BASIC_REC_TFM(BASIC_REC_OUT_TFM), right=MAX_ABPRTY_AGG on EMP_NBR.
   ADJUST_PRTY_TFM: Final AWAY priority calculation → ALL_REC_FNL (via DSLink85).

6. ALL_REC_FNL (PxFunnel)
   Merges all 7 phone-type streams into a single output: ALL_REC_OUT_FNL.

7. NULL_VAL_TFM (CTransformerStage)
   Final null-cleaning and formatting:
   - EMP_NO: left-pads EMP_NBR with zeros to 9 chars
   - EFF_START_TM/EFF_END_TM: formatted as HH:MM:SS
   - PH_ACCESS / PH_COMMENTS: blank/space → NULL
   - TD_LD_TS: CurrentTimestamp()
   Output filter: isvalid('int64', PH_NBR) — invalid records go to WSSEN_ERR_SEQ.

8. TE_EMPLOYEE_PHONE_DW (Target)
   Writes 15 columns to Teradata table CREW_TDB.TE_EMPLOYEE_PHONE.

9. WSSEN_ERR_SEQ (Target)
   Captures error records (invalid EMP_NBR or PH_NBR) to a sequential file.
"""


# ---------------------------------------------------------------------------
# Main build function
# ---------------------------------------------------------------------------
def main():
    print(f"Parsing XML: {XML_PATH}")
    derivations, constraints, _ = parse_xml_derivations(XML_PATH)

    out_links, in_links = build_graph(LINKS)

    # Ensure all stage names appear in the adjacency maps
    for s in STAGE_NAMES_ORDERED:
        if s not in out_links:
            out_links[s] = []
        if s not in in_links:
            in_links[s] = []

    seq_map = assign_sequence_numbers(STAGE_NAMES_ORDERED, out_links, in_links)

    mermaid_str = build_mermaid(STAGE_NAMES_ORDERED, LINKS, out_links, in_links)

    # ------------------------------------------------------------------
    # Build Excel workbook
    # ------------------------------------------------------------------
    wb = openpyxl.Workbook()

    # ---- Tab 1: Stage_Sequence ----
    ws1 = wb.active
    ws1.title = "Stage_Sequence"

    hdrs1 = [
        "Sequence_No", "Stage_Name", "Stage_Type", "Stage_Category",
        "Input_Stages", "Output_Stages", "Parallel_Group",
        "Stage_Description", "Link_Names_In", "Link_Names_Out",
        "Transformation_Summary",
    ]
    ws1.append(hdrs1)

    stage_descriptions = {
        "WSTELE_LND_TDC":      "Teradata source connector reading CREW_WSTELE_LND landing zone table",
        "SEPARATE_PHTYPE_TFM": "Transformer routing records to 6 output streams based on phone type and data validity",
        "BASIC_TYPE_PVT":      "Pivot stage expanding 5 BASIC phone columns per employee into individual rows",
        "BASIC_REC_TFM":       "Transformer mapping pivoted BASIC rows to standard phone record columns",
        "HOME_SEQ1_PVT":       "Pivot stage expanding 5 HOME SEQ1 priority columns per employee into individual rows",
        "HOME_CPY":            "Copy stage duplicating HOME stream into 3 branches for SEQ1/2/3 pivot processing",
        "HOME_SEQ2_PVT":       "Pivot stage expanding 5 HOME SEQ2 priority columns per employee into individual rows; output feeds HOME_SEQ2_TFM",
        "HOME_SEQ3_PVT":       "Pivot stage expanding 5 HOME SEQ3 priority columns per employee into individual rows; output feeds HOME_SEQ3_TFM",
        "HOME_SEQ_FNL":        "Funnel merging HOME SEQ1 pivot + HOME_SEQ2_TFM + HOME_SEQ3_TFM streams into single stream",
        "INC_PRTY_TFM":        "Transformer incrementing CALL_PRIORITY by pivot index to generate 1-based priorities; feeds left-input of HOME_BASIC_JNR",
        "HOME_BASIC_JNR":      "Join: left=INC_PRTY_TFM (HOME pivoted priorities), right=BASIC_REC_TFM (CALL_PRIORITY_RIGHT_JNR) on EMP_NBR; output HBREC_OUT_JNR→NEW_PRTY_TFM",
        "NEW_PRTY_TFM":        "Transformer assigning new call list priorities; output NPRTY_OUT_TFM→ALL_REC_FNL, CPRTY_OUT_TFM→MAX_PRTY_AGG",
        "HOME_SEQ2_TFM":       "Transformer adjusting priorities for HOME SEQ2 records (+5 offset); output feeds HOME_SEQ_FNL",
        "HOME_SEQ3_TFM":       "Transformer adjusting priorities for HOME SEQ3 records (+10 offset); output feeds HOME_SEQ_FNL",
        "MAX_PRTY_AGG":        "Aggregator computing MAX(CALL_PRTY) GROUP BY EMP_NBR for HOME priority records (from NEW_PRTY_TFM)",
        "BASIC_MAX_JNR":       "Join: left=BASIC_REC_TFM (BREC_OUT_TFM), right=MAX_PRTY_AGG (HMAX_OUT_AGG) on EMP_NBR; output→ADJUSTING_PRTY_TFM",
        "ADJUSTING_PRTY_TFM":  "Transformer computing final adjusted PH_PRTY for BASIC records based on max priority join; output ADJ_PRTY_OUT_TFM→ALL_REC_FNL",
        "ALL_REC_FNL":         "Funnel merging all 7 phone type streams into single output stream",
        "AWAY_SEQ3_TFM":       "Transformer adjusting priorities for AWAY SEQ3 records (+10 offset)",
        "AWAY_SEQ2_TFM":       "Transformer adjusting priorities for AWAY SEQ2 records (+5 offset)",
        "AWAY_SEQ_FNL":        "Funnel merging AWAY SEQ1 pivot (ASEQ1_OUT_PVT) + AWAY_SEQ2_TFM + AWAY_SEQ3_TFM streams into single stream",
        "AWAY_SEQ3_PVT":       "Pivot stage expanding 5 AWAY SEQ3 priority columns per employee into individual rows; output feeds AWAY_SEQ3_TFM",
        "AWAY_SEQ2_PVT":       "Pivot stage expanding 5 AWAY SEQ2 priority columns per employee into individual rows; output feeds AWAY_SEQ2_TFM",
        "AWAY_CPY":            "Copy stage duplicating AWAY stream into 3 branches for SEQ1/2/3 pivot processing",
        "AWAY_SEQ1_PVT":       "Pivot stage expanding 5 AWAY SEQ1 priority columns per employee into individual rows; output feeds AWAY_SEQ_FNL directly",
        "INC_AWAY_PRTY_TFM":   "Transformer incrementing AWAY CALL_PRIORITY by pivot index to generate 1-based priorities; feeds left-input of AWAY_BASIC_JNR",
        "AWAY_BASIC_JNR":      "Join: left=INC_AWAY_PRTY_TFM (AWAY pivoted priorities), right=BASIC_REC_TFM (CALL_PRI_RIGHT_JNR) on EMP_NBR; output ABREC_OUT_JNR→ABREC_NEW_PRTY_TFM",
        "MAX_ABPRTY_AGG":      "Aggregator computing MAX(CALL_PRTY) GROUP BY EMP_NBR for AWAY priority records (from ABREC_NEW_PRTY_TFM)",
        "ABREC_NEW_PRTY_TFM":  "Transformer assigning new call priorities to AWAY records; ABNEW_PRTY_OUT_TFM→ALL_REC_FNL, ABCPRTY_OUT_TFM→MAX_ABPRTY_AGG",
        "AWAY_MAX_JNR":        "Join: left=BASIC_REC_TFM (BASIC_REC_OUT_TFM), right=MAX_ABPRTY_AGG (AMAX_OUT_AGG) on EMP_NBR; output DSLink85→ADJUST_PRTY_TFM",
        "ADJUST_PRTY_TFM":     "Transformer computing final adjusted PH_PRTY for AWAY records; output ADJUST_PRTY_OUT_TFM→ALL_REC_FNL",
        "TE_EMPLOYEE_PHONE_DW":"Teradata target connector writing to TE_EMPLOYEE_PHONE warehouse table",
        "NULL_VAL_TFM":        "Final transformer applying null handling, EMP_NO zero-padding, time formatting, and output filter",
        "WSSEN_ERR_SEQ":       "Sequential file target capturing records with invalid EMP_NBR or PH_NBR",
    }

    for name in STAGE_NAMES_ORDERED:
        stype   = STAGE_TYPE_IDS.get(name, "Unknown")
        cat     = get_category(name, stype, in_links, out_links)
        seq     = seq_map.get(name, "")
        in_stgs = "; ".join(sorted(set(src for src, _ in in_links[name])))
        out_stgs= "; ".join(sorted(set(tgt for tgt, _ in out_links[name])))
        lnk_in  = "; ".join(sorted(set(lnk for _, lnk in in_links[name])))
        lnk_out = "; ".join(sorted(set(lnk for _, lnk in out_links[name])))
        par_grp = ""  # could derive from seq number suffix
        if seq and seq[-1].isalpha():
            par_grp = seq[:-1]

        ws1.append([
            seq, name, stype, cat,
            in_stgs, out_stgs, par_grp,
            stage_descriptions.get(name, ""),
            lnk_in, lnk_out,
            get_transformation_summary(name, stype, in_links, out_links),
        ])

    col_widths1 = [12, 30, 22, 16, 35, 35, 14, 45, 35, 35, 50]
    style_sheet(ws1, col_widths1)

    # ---- Tab 2: Mermaid_Lineage ----
    ws2 = wb.create_sheet("Mermaid_Lineage")
    ws2.column_dimensions["A"].width = 120
    ws2.append(["Mermaid Diagram"])
    ws2["A1"].font  = HDR_FONT
    ws2["A1"].fill  = HDR_FILL
    ws2["A1"].alignment = HDR_ALIGN

    ws2.append([mermaid_str])
    ws2["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[2].height = 600

    ws2.append([""])
    ws2.append([LINEAGE_SUMMARY])
    ws2["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[4].height = 400

    # ---- Tab 3: Source_to_Target_Mapping ----
    ws3 = wb.create_sheet("Source_to_Target_Mapping")

    hdrs3 = [
        "Target_Stage", "Target_Table_or_File", "Target_Column",
        "Source_Stage", "Source_Table_or_File", "Source_Column",
        "Stage_Path", "Source_to_Target_Path", "Derivation_Logic",
        "Transformation_Types", "Join_Logic", "Filter_Logic",
        "Lookup_Logic", "Aggregation_Logic", "Default_or_Constant_Logic",
        "Mapping_Type", "Remarks",
    ]
    ws3.append(hdrs3)

    # TE_EMPLOYEE_PHONE_DW rows
    for col_name, info in TARGET_COLUMNS_MAP.items():
        deriv  = info["derivation"]
        src_col = info["source_col"]
        mtype  = info["mapping_type"]
        ttype  = info["trfm_types"]
        path   = STAGE_PATH_MAP.get(col_name, "WSTELE_LND_TDC → ... → NULL_VAL_TFM → TE_EMPLOYEE_PHONE_DW")
        join_l = JOIN_LOGIC_MAP.get(col_name, "")
        filt_l = FILTER_LOGIC_MAP.get(col_name, "NULL_VAL_TFM output filter: isvalid('int64', PH_NBR)")
        agg_l  = AGG_LOGIC_MAP.get(col_name, "")
        const_l = "CurrentTimestamp()" if mtype == "Constant" else ""
        src_tbl = "CREW_WSTELE_LND (Teradata)" if mtype != "Constant" else "N/A"

        ws3.append([
            "TE_EMPLOYEE_PHONE_DW",        # Target_Stage
            "CREW_TDB.TE_EMPLOYEE_PHONE",  # Target_Table_or_File
            col_name,                       # Target_Column
            "NULL_VAL_TFM",                # Source_Stage (immediate)
            src_tbl,                       # Source_Table_or_File
            src_col,                       # Source_Column
            path,                          # Stage_Path
            f"WSTELE_LND_TDC (CREW_WSTELE_LND) → {path.split('→')[-1].strip() if '→' in path else path}",
            deriv,                         # Derivation_Logic
            ttype,                         # Transformation_Types
            join_l,                        # Join_Logic
            filt_l,                        # Filter_Logic
            "",                            # Lookup_Logic
            agg_l,                         # Aggregation_Logic
            const_l,                       # Default_or_Constant_Logic
            mtype,                         # Mapping_Type
            "",                            # Remarks
        ])

    # WSSEN_ERR_SEQ rows
    for col_name in ERROR_COLUMNS:
        ws3.append([
            "WSSEN_ERR_SEQ",               # Target_Stage
            "Sequential error file",        # Target_Table_or_File
            col_name,                       # Target_Column
            "SEPARATE_PHTYPE_TFM",         # Source_Stage
            "CREW_WSTELE_LND (Teradata)",  # Source_Table_or_File
            col_name,                      # Source_Column
            "WSTELE_LND_TDC → SEPARATE_PHTYPE_TFM → WSSEN_ERR_SEQ",
            "WSTELE_LND_TDC (CREW_WSTELE_LND) → SEPARATE_PHTYPE_TFM → WSSEN_ERR_SEQ",
            "LNK_OUT_TDC." + col_name,     # Derivation_Logic
            "Pass-through",                # Transformation_Types
            "",                            # Join_Logic
            "SEPARATE_PHTYPE_TFM: records failing is_valid('INT32',EMP_NBR) AND all phone-type validity checks",
            "",                            # Lookup_Logic
            "",                            # Aggregation_Logic
            "",                            # Default_or_Constant_Logic
            "Pass-through",               # Mapping_Type
            "Error records routed when EMP_NBR is not a valid INT32 or all phone numbers are null/invalid",
        ])

    col_widths3 = [25, 30, 22, 25, 30, 35, 55, 55, 55, 25, 50, 45, 20, 40, 35, 16, 40]
    style_sheet(ws3, col_widths3)

    # ------------------------------------------------------------------
    # Save Excel
    # ------------------------------------------------------------------
    wb.save(XLSX_OUT)
    print(f"Saved Excel: {XLSX_OUT}")

    # ------------------------------------------------------------------
    # Write Mermaid MD file
    # ------------------------------------------------------------------
    md_content = f"# {JOB_NAME}\n\n```mermaid\n{mermaid_str}\n```\n"
    with open(MD_OUT, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Saved Mermaid: {MD_OUT}")

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    unique_links = len(set((s, t, l) for s, t, l in LINKS))
    print("\n=== Summary ===")
    print(f"  Stages found  : {len(STAGE_NAMES_ORDERED)}")
    print(f"  Links found   : {unique_links}")
    print(f"  Target columns (TE_EMPLOYEE_PHONE_DW): {len(TARGET_COLUMNS_MAP)}")
    print(f"  Target columns (WSSEN_ERR_SEQ):        {len(ERROR_COLUMNS)}")
    print(f"  Both output files created: YES")


if __name__ == "__main__":
    main()
